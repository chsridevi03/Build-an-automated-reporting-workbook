import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_mock_data():
    """Generates a sample CSV file if it doesn't exist."""
    data = {
        "Date": pd.date_range(start="2026-01-01", periods=10, freq="D").strftime(
            "%Y-%m-%d"
        ),
        "Region": [
            "North",
            "South",
            "East",
            "West",
            "North",
            "South",
            "East",
            "West",
            "North",
            "South",
        ],
        "Product": [
            "Widget A",
            "Widget B",
            "Widget A",
            "Widget C",
            "Widget B",
            "Widget C",
            "Widget A",
            "Widget B",
            "Widget C",
            "Widget A",
        ],
        "Units Sold": [120, 85, 150, 90, 110, 95, 130, 70, 115, 140],
        "Revenue": [2400, 2125, 3000, 2250, 2750, 2375, 2600, 1750, 2875, 2800],
    }
    df = pd.DataFrame(data)
    df.to_csv("sales_data.csv", index=False)
    print("👉 'sales_data.csv' not found. Generated a new mock data file.")


def create_automated_report(input_file, output_file):
    """Reads raw data, processes it, and builds a formatted Excel Report."""
    print("🔄 Processing data...")

    # 1. Read and aggregate data using Pandas
    df = pd.read_csv(input_file)
    summary_df = (
        df.groupby("Region")[["Units Sold", "Revenue"]].sum().reset_index()
    )

    # 2. Initialize openpyxl workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Regional Performance"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # 3. Design Styling Palette
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=18, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11)
    total_font = Font(name=font_family, size=11, bold=True)

    navy_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    ice_blue_fill = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    total_border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )

    # 4. Add Report Title Block
    ws.merge_cells("A1:C2")
    title_cell = ws["A1"]
    title_cell.value = "Executive Sales Report"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row spacing
    ws.append([])
    ws.append([])

    # 5. Write Dataframe to Excel Sheet
    # Write Headers
    headers = list(summary_df.columns)
    ws.append(headers)
    for cell in ws[5]:
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")

    # Write Data Rows
    for row in dataframe_to_rows(summary_df, index=False, header=False):
        ws.append(row)

    # Style Data Rows
    for row in ws.iter_rows(
        min_row=6, max_row=5 + len(summary_df), min_col=1, max_col=3
    ):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")

            # Format Currency for Revenue Column
            if cell.column == 3:
                cell.number_format = "$#,##0"

    # 6. Add Totals Row
    total_row_idx = 6 + len(summary_df)
    ws.cell(row=total_row_idx, column=1, value="Total").font = total_font

    # Excel Formulas for Totals
    units_total_formula = f"=SUM(B6:B{total_row_idx-1})"
    revenue_total_formula = f"=SUM(C6:C{total_row_idx-1})"

    cell_units = ws.cell(row=total_row_idx, column=2, value=units_total_formula)
    cell_units.font = total_font
    cell_units.number_format = "#,##0"

    cell_rev = ws.cell(row=total_row_idx, column=3, value=revenue_total_formula)
    cell_rev.font = total_font
    cell_rev.number_format = "$#,##0"

    # Apply styling to total row
    for col in range(1, 4):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.border = total_border
        cell.fill = ice_blue_fill

    # 7. Add visual Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Region"
    chart.legend = None

    # References for Data and Categories
    data_ref = Reference(ws, min_col=3, min_row=5, max_row=total_row_idx - 1)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=total_row_idx - 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Position the chart
    ws.add_chart(chart, "E5")

    # 8. Auto-fit column widths dynamically (FIXED BUG HERE)
    for col in ws.columns:
        max_len = 0
        # Dynamically grab the column letter using the first cell's column index
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            # Skip checking the merged banner row values to avoid skewed width sizing
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save the polished workbook
    wb.save(output_file)
    print(f"🚀 Success! Report saved as '{output_file}'")


if __name__ == "__main__":
    input_csv = "sales_data.csv"
    output_excel = "Final_Sales_Report.xlsx"

    # Check for raw data file
    if not os.path.exists(input_csv):
        generate_mock_data()

    # Run automation script
    create_automated_report(input_csv, output_excel)